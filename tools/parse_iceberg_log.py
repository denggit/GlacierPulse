#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Author     : Zijun Deng
@Date       : 5/1/2026 2:50 PM
@File       : parse_iceberg_log.py
@Description: 
"""
import os
import glob
import re
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def parse_iceberg_logs(log_directory="."):
    """解析指定目录下的所有 Phase 1 冰山引擎日志"""

    # 获取所有的日志文件 (包括 app.log 和带有日期的归档日志)
    log_files = glob.glob(os.path.join(log_directory, "app.log*"))

    parsed_data = []
    current_record = {}

    # 预编译正则，提升解析数十万行日志的速度
    re_start = re.compile(r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}).*?CVD: ([\-\d,]+) U.*?触碰价格: ([\d\.]+)")
    re_close = re.compile(
        r"耗时: ([\d\.]+)s.*?战火区: \[([\d\.]+), ([\d\.]+)\].*?总砸盘: ([\d,]+) U.*?盘口消耗: ([\-\d,]+) U")
    re_iceberg = re.compile(r"确信度: ([\d\.]+).*?隐藏体量: ([\d,]+) U.*?吸收率: ([\d\.]+)%")
    re_spoof = re.compile(r"虚假支撑消失量: ([\d,]+) U")

    print(f"找到 {len(log_files)} 个日志文件，开始解析...")

    for file_path in log_files:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                if '[流速点火]' in line:
                    # 如果上一次点火后没有收到确切的冰山或欺诈信号，标记为无信号
                    if current_record:
                        if '结果' not in current_record:
                            current_record['结果'] = '无信号'
                        parsed_data.append(current_record)
                        current_record = {}

                    match = re_start.search(line)
                    if match:
                        current_record['点火时间'] = match.group(1)
                        current_record['触发CVD (U)'] = int(match.group(2).replace(',', ''))
                        current_record['触碰价格'] = float(match.group(3))

                elif '[观测结算]' in line:
                    match = re_close.search(line)
                    if match and current_record:
                        current_record['耗时 (s)'] = float(match.group(1))
                        current_record['战火区底'] = float(match.group(2))
                        current_record['战火区顶'] = float(match.group(3))
                        current_record['总砸盘 (U)'] = int(match.group(4).replace(',', ''))
                        current_record['盘口消耗 (U)'] = int(match.group(5).replace(',', ''))

                elif '[捕获冰山!]' in line:
                    match = re_iceberg.search(line)
                    if match and current_record:
                        current_record['结果'] = '捕获冰山'
                        current_record['确信度'] = float(match.group(1))
                        current_record['隐藏体量/虚假撤单 (U)'] = int(match.group(2).replace(',', ''))
                        current_record['吸收率 (%)'] = float(match.group(3)) / 100.0

                elif '[撤单欺诈!]' in line:
                    match = re_spoof.search(line)
                    if match and current_record:
                        current_record['结果'] = '撤单欺诈'
                        current_record['确信度'] = None
                        current_record['隐藏体量/虚假撤单 (U)'] = int(match.group(1).replace(',', ''))
                        current_record['吸收率 (%)'] = None

    # 处理文件末尾最后一条记录
    if current_record:
        if '结果' not in current_record:
            current_record['结果'] = '无信号'
        parsed_data.append(current_record)

    # 转化为 DataFrame
    df = pd.DataFrame(parsed_data)

    # 确保字段按指定顺序排列
    cols = ['点火时间', '触碰价格', '触发CVD (U)', '耗时 (s)', '战火区底', '战火区顶', '总砸盘 (U)', '盘口消耗 (U)',
            '结果', '确信度', '隐藏体量/虚假撤单 (U)', '吸收率 (%)']
    for col in cols:
        if col not in df.columns:
            df[col] = None

    df = df[cols]

    # 按点火时间排个序
    df = df.sort_values(by='点火时间')

    # 导出到 Excel
    output_file = 'Iceberg_Total_Analysis.xlsx'
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='冰山日志汇总')

    worksheet = writer.sheets['冰山日志汇总']

    # 美化 Excel
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(cols) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        worksheet.column_dimensions[get_column_letter(col)].width = 18

    writer.close()
    print(f"✅ 解析完成！共提取 {len(df)} 次微观探测记录。报表已保存为: {output_file}")


if __name__ == "__main__":
    parse_iceberg_logs(log_directory='../logs/')